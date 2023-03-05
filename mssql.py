import pymysql
from openpyxl import load_workbook


connection = pymysql.connect(
    host='localhost',
    port=3306,
    user='root',
    password='just4Taqtile',
    database='test',
    # charset='utf8mb4',
    cursorclass=pymysql.cursors.DictCursor
)
cursor = connection.cursor()

cursor.execute('''
    CREATE TABLE IF NOT EXISTS Skany (
        Indeks INT PRIMARY KEY,
        Archiwum INT,
        Data DATETIME,
        Del INT,
        KodKreskowy VARCHAR(250),
        Oscieznica INT,
        Pozycja INT,
        Skrzydlo INT,
        srcdoc INT,
        Stanowisko INT,
        Sztuka INT,
        Uzytkownik INT,
        Zakonczony INT,
        Czynnosc INT,
        DbWHOkna INT,
        Guid VARCHAR(250),
        GuidParent VARCHAR(250),
        Status INT,
        Typ INT,
        TypSlupka INT,
        ErrIdx INT
    )
''')

# cursor.execute('''
#     CREATE TABLE IF NOT EXISTS Zlecenia (
#         Indeks INT PRIMARY KEY,
#         Archiwum INT,
#         Data DATETIME,
#         DataWejscia DATETIME,
#         DataZakonczenia DATETIME,
#         Del INT,
#         Diler VARCHAR(250),
#         FirstStanowisko INT,
#         Hiden INT,
#         ErrIdx INT,
#         Klient VARCHAR(250),
#         LiczbaSzklen INT,
#         NipDilera VARCHAR(250),
#         Oscieznica INT,
#         Pozycja INT,
#         Skanowanie INT,
#         Skrzydlo INT,
#         srcdoc INT,
#         Stanowisko INT,
#         StanowiskoPoprzednie INT,
#         Sztuka INT,
#         TerminRealizacji VARCHAR(250),
#         Zakonczone INT,
#         Zlecenie VARCHAR(250),
#         ZlecenieDilera INT,
#         DodOpis VARCHAR(250),
#         optym INT,
#         TerminProdukcji VARCHAR(250),
#         Optymalizacja VARCHAR(250),
#         DbWHOkna INT,
#         KodBiura VARCHAR(250),
#         OptSrcdoc INT,
#         Vip INT,
#         ObrazekOsc VARCHAR(250),
#         ObrazekSkr VARCHAR(250),
#         Referencja VARCHAR(250),
#         Priorytet INT,
#         IloscJedn FLOAT,
#         Idx_typu INT,
#         Typ VARCHAR(250),
#         IloscJednPoz FLOAT,
#         PozycjaLp INT,
#         Country VARCHAR(250),
#         FrameWidth INT,
#         FrameHeight INT,
#         SashWidth INT,
#         SashHeight INT,
#         Glazing VARCHAR(250),
#         GlazingFrame VARCHAR(250),
#         GlazingFrameColor VARCHAR(250),
#         Color VARCHAR(250),
#         Paczka VARCHAR(250)
#     )
# ''')
################################################################
# with connection.cursor() as cursor:
#     create_table_Skany_vs_Zlecenia = "CREATE TABLE `Skany_vs_Zlecenia` (Indeks int, IndeksSkanu int, IndeksZlecenia int, IndeksDodatka varchar(50), Duplicated int);"
#     cursor.execute(create_table_Skany_vs_Zlecenia)
################################################################

# cursor.execute('''
#     CREATE TABLE IF NOT EXISTS Skany_vs_Zlecenia (
#         Indeks INT PRIMARY KEY,
#         IndeksSkanu INT,
#         IndeksZlecenia INT,
#         IndeksDodatka VARCHAR(250),
#         Duplicated INT
#     )
# ''')

#
# cursor.execute('''
#     CREATE TABLE IF NOT EXISTS Stanowiska (
#         Indeks INT PRIMARY KEY,
#         Aktywny INT,
#         Data DATETIME,
#         Del INT,
#         DrukujRaport VARCHAR(250),
#         LiczbaPorzadkowa INT,
#         LiniaProdukcyjna INT,
#         ObslugaStojakow INT,
#         Opis VARCHAR(250),
#         OpisCzynnosci VARCHAR(250),
#         PodstatusPrzed VARCHAR(250),
#         PodstatusPo VARCHAR(250),
#         Raport VARCHAR(250),
#         RaportDodatki VARCHAR(250),
#         RozwinTabelke INT,
#         Skanowanie INT,
#         StanowiskoKoncowe INT,
#         WielkoscCzcionki INT,
#         Zdejmowanie INT,
#         Zliczanie INT,
#         Zoom1 INT,
#         Zoom2 INT,
#         ProceduraSkladowa INT,
#         Viewer VARCHAR(250),
#         CzynnoscOsc INT,
#         CzynnoscSkr INT,
#         CzynnoscSlr INT,
#         CzynnoscSls INT,
#         CzynnoscSzkl INT,
#         ObslugaTransportu INT,
#         BarcodeIdx INT,
#         BarcodePrevIdx INT,
#         BarcodeNextIdx INT,
#         CursorTimeout INT,
#         DefaultEvent INT,
#         TableFilter INT,
#         PanelInfoWidth INT,
#         Printer VARCHAR(250),
#         RaportStojaki VARCHAR(250),
#         ZoomStands INT,
#         Middle INT,
#         Middle_type INT,
#         ObslugaSektorow INT,
#         UserDescription VARCHAR(250),
#         UserStatus VARCHAR(250),
#         CanHaveDifferentIP VARCHAR(250),
#         QualityControlWorkplace VARCHAR(250),
#         AlVARCHAR VARCHAR(250),
#         rasWorkplace INT,
#         AllowGlassScan INT,
#         OnlyOneWorkerOnThisWorkplace INT,
#         AlTEXTrasDateColumnName VARCHAR(250),
#         HideLaborButton INT,
#         ImportPackagesToSzybyXLS INT,
#         HideTableInPackagesLoading INT,
#         AltCuttingWorkplace INT,
#         Mobile INT,
#         markwhentransportispacked INT
#     )
# ''')
#
# cursor.execute('''
#     CREATE TABLE IF NOT EXISTS Uzytkownicy (
#         Indeks INT PRIMARY KEY,
#         Aktywny INT,
#         Data DATETIME,
#         Dealer VARCHAR(250),
#         Del INT,
#         Haslo VARCHAR(250),
#         Imie VARCHAR(250),
#         Login VARCHAR(250),
#         Nazwa VARCHAR(250),
#         Nazwisko VARCHAR(250),
#         Nip VARCHAR(250),
#         Uprawnienia INT,
#         Usr INT,
#         Uwagi VARCHAR(250),
#         StawkaDzienna INT,
#         BarcodeIdx INT,
#         Language VARCHAR(250),
#         GrupaPlacowa INT,
#         TworzenieArtykulow INT,
#         Email VARCHAR(250),
#         ZestawienieZlecenNaProdukcjiVisible INT,
#         ZawartoscStojakowVisible INT,
#         ZawartoscSamochodowVisible INT,
#         ZawartoscSektorowVisible INT,
#         ZawartoscSektorowSzkleniaVisible INT,
#         ZestawienieCzynnosciVisible INT,
#         ZestawienieOdpowiedziNaPytaniaVisible INT,
#         ZestawienieRobociznyVisible INT,
#         ZestawienieBledowKomunikatowNotatekVisible INT,
#         EksportWykonanychOscieznicVisible INT,
#         PostepRealizacjiVisible INT,
#         DodajPracownikaVisible INT,
#         CofnijSkanVisible INT,
#         ZestawienieCzynnosciNewVisible INT,
#         VisibilityLastDateChange VARCHAR(250),
#         Image VARCHAR(250)
#     )
# ''')

wb = load_workbook(filename='file/skany.xlsx', read_only=True)
ws = wb.active

sheet_names = wb.sheetnames
worksheets = [wb[sheet_name] for sheet_name in sheet_names]

for worksheet in worksheets:
    i = 0
    for row in worksheet.rows:
        if i != 0:
            Indeks = int(row[0].value)
            Archiwum = int(row[1].value)
            Data = row[2].value
            Del = int(row[3].value)
            KodKreskowy = row[4].value
            Oscieznica = int(row[5].value)
            Pozycja = int(row[6].value)
            Skrzydlo = int(row[7].value)
            srcdoc = int(row[8].value)
            Stanowisko = int(row[9].value)
            Sztuka = int(row[10].value)
            Uzytkownik = int(row[11].value)
            Zakonczony = int(row[12].value)
            Czynnosc = int(row[13].value)
            DbWHOkna = int(row[14].value)
            Guid = row[15].value
            GuidParent = row[16].value
            Status = int(row[17].value)
            Typ = int(row[18].value)
            TypSlupka = int(row[19].value)
            ErrIdx = int(row[20].value)
            # print(Indeks, Archiwum, Data, Del, KodKreskowy, Oscieznica, Pozycja, Skrzydlo, srcdoc, Stanowisko, Sztuka, Uzytkownik, Zakonczony, Czynnosc, DbWHOkna, Guid, GuidParent, Status, Typ, TypSlupka, ErrIdx)
            cursor.execute(
                "INSERT INTO Skany(Indeks, Archiwum, Data, Del, KodKreskowy, Oscieznica, Pozycja, Skrzydlo, srcdoc, Stanowisko, Sztuka, Uzytkownik, Zakonczony, Czynnosc, DbWHOkna, Guid, GuidParent, Status, Typ, TypSlupka, ErrIdx) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (Indeks, Archiwum, Data, Del, KodKreskowy, Oscieznica, Pozycja, Skrzydlo, srcdoc, Stanowisko, Sztuka, Uzytkownik, Zakonczony, Czynnosc, DbWHOkna, Guid, GuidParent, Status, Typ, TypSlupka, ErrIdx))
            connection.commit()
        i += 1

# wbz = load_workbook(filename='file/Zlecenia.xlsx', read_only=True)
# wsz = wbz.active
# for row in wsz.rows:
#     Indeks = row[0].value
#     Archiwum = row[1].value
#     Data = row[2].value
#     DataWejscia = row[3].value
#     DataZakonczenia = row[4].value
#     Del = row[5].value
#     Diler = row[6].value
#     FirstStanowisko = row[7].value
#     Hiden = row[8].value
#     ErrIdx = row[9].value
#     Klient = row[10].value
#     LiczbaSzklen = row[11].value
#     NipDilera = row[12].value
#     Oscieznica = row[13].value
#     Pozycja = row[14].value
#     Skanowanie = row[15].value
#     Skrzydlo = row[16].value
#     srcdoc = row[17].value
#     Stanowisko = row[18].value
#     StanowiskoPoprzednie = row[19].value
#     Sztuka = row[20].value
#     TerminRealizacji = row[21].value
#     Zakonczone = row[22].value
#     Zlecenie = row[23].value
#     ZlecenieDilera = row[24].value
#     DodOpis = row[25].value
#     optym = row[26].value
#     TerminProdukcji = row[27].value
#     Optymalizacja = row[28].value
#     DbWHOkna = row[29].value
#     KodBiura = row[30].value
#     OptSrcdoc = row[31].value
#     Vip = row[32].value
#     ObrazekOsc = row[33].value
#     ObrazekSkr = row[34].value
#     Referencja = row[35].value
#     Priorytet = row[36].value
#     IloscJedn = row[37].value
#     Idx_typu = row[38].value
#     Typ = row[39].value
#     IloscJednPoz = row[40].value
#     PozycjaLp = row[41].value
#     Country = row[42].value
#     FrameWidth = row[43].value
#     FrameHeight = row[44].value
#     SashWidth = row[45].value
#     SashHeight = row[46].value
#     Glazing = row[47].value
#     GlazingFrame = row[48].value
#     GlazingFrameColor = row[49].value
#     Color = row[50].value
#     Paczka = row[51].value
#
#     cursor.execute(
#         "INSERT INTO Zlecenia (Indeks, Archiwum, Data, DataWejscia, DataZakonczenia, Del, Diler, FirstStanowisko, Hiden, ErrIdx, Klient, LiczbaSzklen, NipDilera, Oscieznica, Pozycja, Skanowanie, Skrzydlo, srcdoc, Stanowisko, StanowiskoPoprzednie, Sztuka, TerminRealizacji, Zakonczone, Zlecenie, ZlecenieDilera, DodOpis, optym, TerminProdukcji, Optymalizacja, DbWHOkna, KodBiura, OptSrcdoc, Vip, ObrazekOsc, ObrazekSkr, Referencja, Priorytet, IloscJedn, Idx_typu, Typ, IloscJednPoz, PozycjaLp, Country, FrameWidth, FrameHeight, SashWidth, SashHeight, Glazing, GlazingFrame, GlazingFrameColor, Color, Paczka) VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) ",
#         (Indeks, Archiwum, Data, DataWejscia, DataZakonczenia, Del, Diler, FirstStanowisko, Hiden, ErrIdx, Klient,
#          LiczbaSzklen, NipDilera, Oscieznica, Pozycja, Skanowanie, Skrzydlo, srcdoc, Stanowisko,
#          StanowiskoPoprzednie, Sztuka, TerminRealizacji, Zakonczone, Zlecenie, ZlecenieDilera, DodOpis, optym,
#          TerminProdukcji, Optymalizacja, DbWHOkna, KodBiura, OptSrcdoc, Vip, ObrazekOsc, ObrazekSkr, Referencja,
#          Priorytet, IloscJedn, Idx_typu, Typ, IloscJednPoz, PozycjaLp, Country, FrameWidth, FrameHeight, SashWidth,
#          SashHeight, Glazing, GlazingFrame, GlazingFrameColor, Color, Paczka))
################################################################
# with connection.cursor() as cursor:
#     wbzvz = load_workbook(filename='file/Skany_vs_Zlecenia.xlsx', read_only=True)
#     wszvz = wbzvz.active
#     i = 0
#     for row in wszvz.rows:
#         if i == 0:
#             Indeks = row[0].value
#             IndeksSkanu = row[1].value
#             IndeksZlecenia = row[2].value
#             IndeksDodatka = row[3].value
#             Duplicated = row[4].value
#         else:
#             Indeks = int(row[0].value)
#             IndeksSkanu = int(row[1].value)
#             IndeksZlecenia = int(row[2].value)
#             IndeksDodatka = row[3].value
#             Duplicated = int(row[4].value)
#         # print(Indeks, IndeksSkanu, IndeksZlecenia, IndeksDodatka, Duplicated)
#         insert_data = "INSERT INTO Skany_vs_Zlecenia(Indeks, IndeksSkanu, IndeksZlecenia, IndeksDodatka, Duplicated) VALUES (Indeks, IndeksSkanu, IndeksZlecenia, IndeksDodatka, Duplicated)"
#         cursor.execute(insert_data)
#         connection.commit()
#         i += 1

################################################################
# wbzvz = load_workbook(filename='file/Skany_vs_Zlecenia.xlsx', read_only=True)
# wszvz = wbzvz.active
# i = 0
# for row in wszvz.rows:
#     if i == 0:
#         Indeks = row[0].value
#         IndeksSkanu = row[1].value
#         IndeksZlecenia = row[2].value
#         IndeksDodatka = row[3].value
#         Duplicated = row[4].value
#     else:
#         Indeks = int(row[0].value)
#         IndeksSkanu = int(row[1].value)
#         IndeksZlecenia = int(row[2].value)
#         IndeksDodatka = row[3].value
#         Duplicated = int(row[4].value)
#     # print(Indeks, IndeksSkanu, IndeksZlecenia, IndeksDodatka, Duplicated)
#         cursor.execute(
#             "INSERT INTO Skany_vs_Zlecenia(Indeks, IndeksSkanu, IndeksZlecenia, IndeksDodatka, Duplicated) VALUES (%s, %s, %s, %s, %s)",
#             (Indeks, IndeksSkanu, IndeksZlecenia, IndeksDodatka, Duplicated))
#         connection.commit()
#     i += 1

#
# wbzvzs = load_workbook(filename='file/Stanowiska.xlsx', read_only=True)
# wszvzs = wbzvzs.active
# for row in wszvzs.rows:
#     Indeks = row[0].value
#     Aktywny = row[1].value
#     Data = row[2].value
#     Del = row[3].value
#     DrukujRaport = row[4].value
#     LiczbaPorzadkowa = row[5].value
#     LiniaProdukcyjna = row[6].value
#     ObslugaStojakow = row[7].value
#     Opis = row[8].value
#     OpisCzynnosci = row[9].value
#     PodstatusPrzed = row[10].value
#     PodstatusPo = row[11].value
#     Raport = row[12].value
#     RaportDodatki = row[13].value
#     RozwinTabelke = row[14].value
#     Skanowanie = row[15].value
#     StanowiskoKoncowe = row[16].value
#     WielkoscCzcionki = row[17].value
#     Zdejmowanie = row[18].value
#     Zliczanie = row[19].value
#     Zoom1 = row[20].value
#     Zoom2 = row[21].value
#     ProceduraSkladowa = row[22].value
#     Viewer = row[23].value
#     CzynnoscOsc = row[24].value
#     CzynnoscSkr = row[25].value
#     CzynnoscSlr = row[26].value
#     CzynnoscSls = row[27].value
#     CzynnoscSzkl = row[28].value
#     ObslugaTransportu = row[29].value
#     BarcodeIdx = row[30].value
#     BarcodePrevIdx = row[31].value
#     BarcodeNextIdx = row[32].value
#     CursorTimeout = row[33].value
#     DefaultEvent = row[34].value
#     TableFilter = row[35].value
#     PanelInfoWidth = row[36].value
#     Printer = row[37].value
#     RaportStojaki = row[38].value
#     ZoomStands = row[39].value
#     Middle = row[40].value
#     Middle_type = row[41].value
#     ObslugaSektorow = row[42].value
#     UserDescription = row[43].value
#     UserStatus = row[44].value
#     CanHaveDifferentIP = row[45].value
#     QualityControlWorkplace = row[46].value
#     AlTEXTrasWorkplace = row[47].value
#     AllowGlassScan = row[48].value
#     OnlyOneWorkerOnThisWorkplace = row[49].value
#     AlTEXTrasDateColumnName = row[50].value
#     HideLaborButton = row[51].value
#     ImportPackagesToSzybyXLS = row[52].value
#     HideTableInPackagesLoading = row[53].value
#     Mobile = row[54].value
#     AltCuttingWorkplace = row[55].value
#     markwhentransportispacked = row[56].value
# print(Indeks, Aktywny, Data, Del, DrukujRaport, LiczbaPorzadkowa, LiniaProdukcyjna, ObslugaStojakow, Opis, OpisCzynnosci, PodstatusPrzed, PodstatusPo, Raport, RaportDodatki, RozwinTabelke, Skanowanie, StanowiskoKoncowe, WielkoscCzcionki, Zdejmowanie, Zliczanie, Zoom1, Zoom2, ProceduraSkladowa, Viewer, CzynnoscOsc, CzynnoscSkr, CzynnoscSlr, CzynnoscSls, CzynnoscSzkl, ObslugaTransportu, BarcodeIdx, BarcodePrevIdx, BarcodeNextIdx, CursorTimeout, DefaultEvent, TableFilter, PanelInfoWidth, Printer, RaportStojaki, ZoomStands, Middle, Middle_type, ObslugaSektorow, UserDescription, UserStatus, CanHaveDifferentIP, QualityControlWorkplace, AlTEXTrasWorkplace, AllowGlassScan, OnlyOneWorkerOnThisWorkplace, AlTEXTrasDateColumnName, HideLaborButton, ImportPackagesToSzybyXLS, HideTableInPackagesLoading, Mobile, AltCuttingWorkplace, markwhentransportispacked)
#     cursor.execute(
#         "INSERT INTO Stanowiska(Indeks, Aktywny, Data, Del, DrukujRaport, LiczbaPorzadkowa, LiniaProdukcyjna, ObslugaStojakow, Opis, OpisCzynnosci, PodstatusPrzed, PodstatusPo, Raport, RaportDodatki, RozwinTabelke, Skanowanie, StanowiskoKoncowe, WielkoscCzcionki, Zdejmowanie, Zliczanie, Zoom1, Zoom2, ProceduraSkladowa, Viewer, CzynnoscOsc, CzynnoscSkr, CzynnoscSlr, CzynnoscSls, CzynnoscSzkl, ObslugaTransportu, BarcodeIdx, BarcodePrevIdx, BarcodeNextIdx, CursorTimeout, DefaultEvent, TableFilter, PanelInfoWidth, Printer, RaportStojaki, ZoomStands, Middle, Middle_type, ObslugaSektorow, UserDescription, UserStatus, CanHaveDifferentIP, QualityControlWorkplace, AlTEXTrasWorkplace, AllowGlassScan, OnlyOneWorkerOnThisWorkplace, AlTEXTrasDateColumnName, HideLaborButton, ImportPackagesToSzybyXLS, HideTableInPackagesLoading, Mobile, AltCuttingWorkplace, markwhentransportispacked) VALUES (%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s)",
#         (Indeks, Aktywny, Data, Del, DrukujRaport, LiczbaPorzadkowa, LiniaProdukcyjna, ObslugaStojakow, Opis,
#          OpisCzynnosci, PodstatusPrzed, PodstatusPo, Raport, RaportDodatki, RozwinTabelke, Skanowanie,
#          StanowiskoKoncowe, WielkoscCzcionki, Zdejmowanie, Zliczanie, Zoom1, Zoom2, ProceduraSkladowa, Viewer,
#          CzynnoscOsc, CzynnoscSkr, CzynnoscSlr, CzynnoscSls, CzynnoscSzkl, ObslugaTransportu, BarcodeIdx,
#          BarcodePrevIdx, BarcodeNextIdx, CursorTimeout, DefaultEvent, TableFilter, PanelInfoWidth, Printer,
#          RaportStojaki, ZoomStands, Middle, Middle_type, ObslugaSektorow, UserDescription, UserStatus,
#          CanHaveDifferentIP, QualityControlWorkplace, AlTEXTrasWorkplace, AllowGlassScan, OnlyOneWorkerOnThisWorkplace,
#          AlTEXTrasDateColumnName, HideLaborButton, ImportPackagesToSzybyXLS, HideTableInPackagesLoading, Mobile,
#          AltCuttingWorkplace, markwhentransportispacked))
#
# #
# wbzvzuz = load_workbook(filename='file/Uzytkownicy.xlsx', read_only=True)
# wszvzu = wbzvzuz.worksheets[0]
# for row in wszvzu.rows:
#     if len(row) == 36:
#         Indeks = row[0].value
#         Aktywny = row[1].value
#         Data = row[2].value
#         Dealer = row[3].value
#         Del = row[4].value
#         Haslo = row[5].value
#         Imie = row[6].value
#         Login = row[7].value
#         Nazwa = row[8].value
#         Nazwisko = row[9].value
#         Nip = row[10].value
#         Uprawnienia = row[11].value
#         Usr = row[12].value
#         Uwagi = row[13].value
#         StawkaDzienna = row[14].value
#         BarcodeIdx = row[15].value
#         Language = row[16].value
#         GrupaPlacowa = row[17].value
#         TworzenieArtykulow = row[18].value
#         Email = row[19].value
#         ZestawienieZlecenNaProdukcjiVisible = row[20].value
#         ZawartoscStojakowVisible = row[21].value
#         ZawartoscSamochodowVisible = row[22].value
#         ZawartoscSektorowVisible = row[23].value
#         ZawartoscSektorowSzkleniaVisible = row[24].value
#         ZestawienieCzynnosciVisible = row[25].value
#         ZestawienieOdpowiedziNaPytaniaVisible = row[26].value
#         ZestawienieRobociznyVisible = row[27].value
#         ZestawienieBledowKomunikatowNotatekVisible = row[28].value
#         EksportWykonanychOscieznicVisible = row[29].value
#         PostepRealizacjiVisible = row[30].value
#         DodajPracownikaVisible = row[31].value
#         CofnijSkanVisible = row[32].value
#         ZestawienieCzynnosciNewVisible = row[33].value
#         VisibilityLastDateChange = row[34].value
#         Image = row[35].value
# print(Indeks, Aktywny, Data, Dealer, Del, Haslo, Imie, Login, Nazwa, Nazwisko, Nip, Uprawnienia, Usr, Uwagi, StawkaDzienna, BarcodeIdx, Language, GrupaPlacowa, TworzenieArtykulow, Email, ZestawienieZlecenNaProdukcjiVisible, ZawartoscStojakowVisible, ZawartoscSamochodowVisible, ZawartoscSektorowVisible, ZawartoscSektorowSzkleniaVisible, ZestawienieCzynnosciVisible, ZestawienieOdpowiedziNaPytaniaVisible, ZestawienieRobociznyVisible, ZestawienieBledowKomunikatowNotatekVisible, EksportWykonanychOscieznicVisible, PostepRealizacjiVisible, DodajPracownikaVisible, CofnijSkanVisible, ZestawienieCzynnosciNewVisible, VisibilityLastDateChange, Image)
# cursor.execute(
#     "INSERT INTO Uzytkownicy(Indeks, Aktywny, Data, Dealer, Del, Haslo, Imie, Login, Nazwa, Nazwisko, Nip, Uprawnienia, Usr, Uwagi, StawkaDzienna, BarcodeIdx, Language, GrupaPlacowa, TworzenieArtykulow, Email, ZestawienieZlecenNaProdukcjiVisible, ZawartoscStojakowVisible, ZawartoscSamochodowVisible, ZawartoscSektorowVisible, ZawartoscSektorowSzkleniaVisible, ZestawienieCzynnosciVisible, ZestawienieOdpowiedziNaPytaniaVisible, ZestawienieRobociznyVisible, ZestawienieBledowKomunikatowNotatekVisible, EksportWykonanychOscieznicVisible, PostepRealizacjiVisible, DodajPracownikaVisible, CofnijSkanVisible, ZestawienieCzynnosciNewVisible, VisibilityLastDateChange, Image) VALUES (%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s, %s, %s, %s, %s ,%s, %s, %s, %s)",
#     (Indeks, Aktywny, Data, Dealer, Del, Haslo, Imie, Login, Nazwa, Nazwisko, Nip, Uprawnienia, Usr, Uwagi,
#      StawkaDzienna, BarcodeIdx, Language, GrupaPlacowa, TworzenieArtykulow, Email,
#      ZestawienieZlecenNaProdukcjiVisible, ZawartoscStojakowVisible, ZawartoscSamochodowVisible,
#      ZawartoscSektorowVisible, ZawartoscSektorowSzkleniaVisible, ZestawienieCzynnosciVisible,
#      ZestawienieOdpowiedziNaPytaniaVisible, ZestawienieRobociznyVisible,
#      ZestawienieBledowKomunikatowNotatekVisible, EksportWykonanychOscieznicVisible, PostepRealizacjiVisible,
#      DodajPracownikaVisible, CofnijSkanVisible, ZestawienieCzynnosciNewVisible, VisibilityLastDateChange,
#      Image))
connection.close()